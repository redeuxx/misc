import os.path
import sys
from openpyxl import load_workbook


def main():
    # main function

    # check to see if file input is valid
    try:
        filename1 = sys.argv[1]
        filename2 = sys.argv[2]
        append = sys.argv[3]
        if not os.path.exists(filename1):
            sys.exit("Invalid file or files.")
        if not os.path.exists(filename2):
            sys.exit("Invalid file or files.")
    except (FileNotFoundError, IndexError):
        sys.exit("Invalid file or files.")

    workbook1 = load_workbook(filename=filename1)
    duo_list = []

    # read first Excel worksheet and put it into a list
    logins_list = []
    sheet1 = workbook1['unique users']
    for cell1 in sheet1["A"]:
        logins_list.append(cell1.value)
    logins_list_sanitized = []
    a = add_names(logins_list, logins_list_sanitized)
    print(len(a))

    # read second Excel worksheet and put it into a list (not currently used)
    logins_list = []  # reset variable
    sheet2 = workbook1['win logins']
    for cell1 in sheet2["C"]:
        logins_list.append(cell1.value)
    logins_list_sanitized = []  # reset variable
    b = add_names(logins_list, logins_list_sanitized)
    print(len(b))

    # remove duplicates from first and second worksheet (not currently used)
    c = add_names(a, b)
    print(len(c))

    # remove duplicates already in Duo
    workbook2 = load_workbook(filename=filename2)
    sheet4 = workbook2.active
    for cell2 in sheet4["A"]:
        duo_list.append(cell2.value)
    e = remove_dupes(a, duo_list)
    print(len(e))
    write_to_file(a, append)


def add_names(list1, list2):
    # creates a list from a file and returns list with no duplicates
    for value in list1:
        if value.lower() not in list2:
            list2.append(value.lower())
    return list2


def remove_dupes(list1, list2):
    # removes usernames already in list 2 from list 1
    for name in list2:
        if name in list1:
            list1.remove(name)
    return list1


def write_to_file(list1, append):
    content = ""
    for value in list1:
        full = value + append
        content = content + "\n" + full
    with open("usernames1.txt", "w") as file1:
        file1.write(content)


if __name__ == '__main__':
    main()
